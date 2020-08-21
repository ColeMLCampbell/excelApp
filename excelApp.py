import openpyxl
from os import path

def load_workbook(wb_path):
    if path.exists(wb_path):
        return openpyxl.load_workbook(wb_path)
    return openpyxl.Workbook()

wb_path = 'test.xlsx'
wb = load_workbook(wb_path)
sheet = wb['Customer Count']

custCount = [
    (1, 'One'),
    (2, 'Two'),
    (3, 'Three')
] 

#for cust in custCount:
#    sheet.append(cust)

#sheet.insert_rows(1, 3)

# for 

 
wb.save(wb_path)

print('Hello is it me youre looking for?')